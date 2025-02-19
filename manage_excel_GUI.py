#!/usr/bin/env python3
"""
Manage Excel Sheets and Files Utility - GUI

This script allows you to combine multiple Excel files into one,
or split a single Excel file into multiple sheets or files based on a specific column.
It includes enhanced handling for password-protected files with a popup password prompt,
caching the unprotected copy to avoid asking the user repeatedly.

Version = "0.2"
Author: Obaid Aldosari
GitHub: https://github.com/ODosari/manage_excel_sheets_and_files
"""

import os
import time
import glob
import logging
import tempfile
import io
import traceback
from contextlib import closing

import pandas as pd
import msoffcrypto
from openpyxl import load_workbook
from msoffcrypto.exceptions import InvalidKeyError

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# --- Logging configuration ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    filename="manage_excel.log",
    filemode="a",
)
logger = logging.getLogger()

# A global cache mapping original_file_path -> temp_unprotected_file_path
# so we don't repeatedly ask for a password on the same file.
UNPROTECTED_CACHE = {}


def get_timestamped_filename(base_path, prefix, extension):
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    return os.path.join(base_path, f"{prefix}_{timestamp}{extension}")


def normalize_sheet_name(base_name, sheet_name, existing_names):
    raw_name = f"{base_name}_{sheet_name}".replace(" ", "_")
    safe_name = raw_name[:31]
    counter = 1
    while safe_name in existing_names:
        suffix = f"_{counter}"
        safe_name = raw_name[:31 - len(suffix)] + suffix
        counter += 1
    existing_names.add(safe_name)
    return safe_name


def combine_files(selected_files, combine_mode="one_sheet"):
    """
    Combine selected Excel files into one, either as a single sheet or multiple sheets.
    """
    if not selected_files:
        return None

    out_dir = os.path.dirname(selected_files[0])
    output_file = get_timestamped_filename(out_dir, "Combined", ".xlsx")

    try:
        if combine_mode == "one_sheet":
            combined_df = pd.DataFrame()
            for file_path in selected_files:
                sheets = get_sheets_from_file(file_path)
                if sheets:
                    df = read_sheet_from_file(file_path, sheets[0])
                    if df is not None:
                        combined_df = pd.concat([combined_df, df], ignore_index=True)
            combined_df.to_excel(output_file, sheet_name="Combined", engine="openpyxl", index=False)
        else:
            base_names_used = set()
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for file_path in selected_files:
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    sheets = get_sheets_from_file(file_path)
                    if sheets:
                        df = read_sheet_from_file(file_path, sheets[0])
                        if df is not None:
                            safe_sheet = normalize_sheet_name(base_name, sheets[0], base_names_used)
                            df.to_excel(writer, sheet_name=safe_sheet, index=False)
        logger.info(f"Files combined successfully. Output File: {os.path.basename(output_file)}")
        return output_file
    except Exception as e:
        logger.error(f"An error occurred while combining files: {e}")
        traceback.print_exc()
        return None


def split_file(file_path, sheet_name, column, split_mode="files"):
    """
    Split an Excel file based on a column's unique values, either into separate files or sheets.
    """
    df = read_sheet_from_file(file_path, sheet_name)
    if df is None:
        return None

    columns = list(df.columns)
    if isinstance(column, int):
        if column < 0 or column >= len(columns):
            logger.error("Invalid column index selected.")
            return None
        split_column = columns[column]
    elif isinstance(column, str):
        if column not in columns:
            logger.error("Invalid column name selected.")
            return None
        split_column = column
    else:
        logger.error("Column must be an int or str.")
        return None

    out_dir = os.path.dirname(file_path)
    try:
        if split_mode == "files":
            output_files = []
            for value in df[split_column].unique():
                safe_value = str(value).replace(" ", "_")
                output_file = get_timestamped_filename(
                    out_dir,
                    f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}_{split_column}_{safe_value}",
                    ".xlsx"
                )
                filtered_df = df[df[split_column] == value]
                filtered_df.to_excel(output_file, sheet_name=str(value)[:31], engine="openpyxl", index=False)
                output_files.append(output_file)
                logger.info(f"Data for '{value}' written to {os.path.basename(output_file)}")
            logger.info("Data split into files successfully.")
            return output_files
        else:
            output_file = get_timestamped_filename(
                out_dir,
                f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}_split",
                ".xlsx"
            )
            base_names_used = set()
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                for value in df[split_column].unique():
                    raw_sheet_name = str(value)
                    safe_sheet = normalize_sheet_name(raw_sheet_name, "", base_names_used) if raw_sheet_name.strip() else "Empty"
                    filtered_df = df[df[split_column] == value]
                    filtered_df.to_excel(writer, sheet_name=safe_sheet, index=False)
            logger.info(f"Data split into sheets successfully. Output File: {os.path.basename(output_file)}")
            return output_file
    except Exception as e:
        logger.error(f"An error occurred while splitting the file: {e}")
        traceback.print_exc()
        return None


def get_sheets_from_file(file_path):
    """
    Return the list of sheet names for an Excel file, prompting for password if needed,
    but only once per session.
    """
    temp_file = unprotect_excel_file_with_prompt(file_path)
    if temp_file is None:
        return []
    try:
        with closing(pd.ExcelFile(temp_file, engine="openpyxl")) as workbook:
            sheets = workbook.sheet_names
        return sheets
    except Exception as e:
        logger.error(f"Error reading sheets from {os.path.basename(file_path)}: {e}")
        return []
    finally:
        # We do NOT delete the temp_file here if we want to reuse it.
        pass


def read_sheet_from_file(file_path, sheet_name):
    """
    Read a single sheet from an Excel file, prompting for password if needed,
    but only once per session.
    """
    temp_file = unprotect_excel_file_with_prompt(file_path)
    if temp_file is None:
        return None
    try:
        df = pd.read_excel(temp_file, sheet_name=sheet_name, engine="openpyxl")
        return df
    except Exception as e:
        logger.error(f"Error reading sheet {sheet_name} from {os.path.basename(file_path)}: {e}")
        return None
    finally:
        # We do NOT delete the temp_file here if we want to reuse it for multiple reads.
        pass


def unprotect_excel_file_with_prompt(file_path, max_attempts=3):
    """
    Returns a path to an unprotected copy of 'file_path', using a cached temp file if available.
    If the file is encrypted, prompts for a password. Only prompts once per session.
    """
    # 1) Check if we already have a cached unprotected file
    if file_path in UNPROTECTED_CACHE:
        return UNPROTECTED_CACHE[file_path]

    try:
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            if office_file.is_encrypted():
                attempts = 0
                while attempts < max_attempts:
                    password = prompt_for_password(os.path.basename(file_path))
                    if password is None:  # user canceled
                        logger.info(f"User canceled password entry for {file_path}")
                        return None
                    try:
                        decrypted = io.BytesIO()
                        office_file.load_key(password=password)
                        office_file.decrypt(decrypted)
                        decrypted.seek(0)
                        wb = load_workbook(decrypted, read_only=False, keep_vba=True)
                        logger.info(f"Password accepted for {os.path.basename(file_path)}")

                        # Remove protection
                        wb.security = None
                        for sheet in wb.worksheets:
                            sheet.protection.enabled = False
                            sheet.protection.sheet = False

                        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                        wb.save(temp_file.name)
                        temp_file.close()
                        logger.debug(f"Unprotected file saved to temp: {temp_file.name}")

                        # Cache it so we don't ask again
                        UNPROTECTED_CACHE[file_path] = temp_file.name
                        return temp_file.name
                    except InvalidKeyError as ike:
                        attempts += 1
                        logger.error(
                            f"Incorrect password for {os.path.basename(file_path)} "
                            f"(Attempt {attempts}/{max_attempts}): {ike}"
                        )
                        if attempts >= max_attempts:
                            messagebox.showerror("Error", "Maximum password attempts reached. Skipping file.")
                            return None
            else:
                # Not encrypted
                wb = load_workbook(file_path, read_only=False, keep_vba=True)
                wb.security = None
                for sheet in wb.worksheets:
                    sheet.protection.enabled = False
                    sheet.protection.sheet = False

                temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                wb.save(temp_file.name)
                temp_file.close()
                logger.debug(f"Unprotected file saved to temp: {temp_file.name}")

                # Cache
                UNPROTECTED_CACHE[file_path] = temp_file.name
                return temp_file.name

    except Exception as e:
        logger.error(f"Failed to open {os.path.basename(file_path)}: {e}")
        traceback.print_exc()
        return None


def prompt_for_password(file_name):
    """
    Shows a popup window asking the user to enter a password for the given file.
    Returns the password as a string, or None if the user cancels.
    """
    popup = tk.Toplevel()
    popup.title(f"Enter Password - {file_name}")
    popup.configure(bg="#2e2e2e")

    # Make the window modal
    popup.grab_set()
    popup.resizable(False, False)

    label = ttk.Label(popup, text=f"File '{file_name}' is password-protected.\nEnter password:")
    label.pack(padx=10, pady=10)

    pwd_var = tk.StringVar()

    entry = ttk.Entry(popup, textvariable=pwd_var, show="*", width=30)
    entry.pack(padx=10, pady=5)
    entry.focus()

    result = [None]

    def on_ok():
        result[0] = pwd_var.get()
        popup.destroy()

    def on_cancel():
        result[0] = None
        popup.destroy()

    btn_frame = ttk.Frame(popup, style="TFrame")
    btn_frame.pack(pady=10)
    ok_btn = ttk.Button(btn_frame, text="OK", command=on_ok)
    ok_btn.pack(side=tk.LEFT, padx=5)
    cancel_btn = ttk.Button(btn_frame, text="Cancel", command=on_cancel)
    cancel_btn.pack(side=tk.LEFT, padx=5)

    popup.bind("<Return>", lambda e: on_ok())
    popup.bind("<Escape>", lambda e: on_cancel())

    # Center popup
    popup.update_idletasks()
    w = 400
    h = 200
    x = (popup.winfo_screenwidth() // 2) - (w // 2)
    y = (popup.winfo_screenheight() // 2) - (h // 2)
    popup.geometry(f"{w}x{h}+{x}+{y}")

    popup.wait_window()
    return result[0]


class ExcelManagerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Manage Excel Sheets and Files")
        self.geometry("1000x700")
        self.center_window()

        # Dark theme
        self.style = ttk.Style(self)
        self.style.theme_use("clam")
        self.configure(bg="#2e2e2e")

        # Notebook styling
        self.style.configure("TNotebook", background="#2e2e2e", borderwidth=0)
        self.style.configure("TNotebook.Tab", background="#2e2e2e", foreground="#ffffff", padding=[10, 5])
        self.style.map("TNotebook.Tab",
                       background=[("selected", "#00a67d"), ("active", "#00a67d")],
                       foreground=[("selected", "#ffffff"), ("active", "#ffffff")]
                       )

        self.style.configure("TLabelframe", background="#2e2e2e", borderwidth=0)
        self.style.configure("TLabelframe.Label", background="#2e2e2e", foreground="#ffffff")

        self.style.configure(".", font=("IBM Plex Mono", 16))
        self.style.configure("TFrame", background="#2e2e2e")
        self.style.configure("TLabel", background="#2e2e2e", foreground="#ffffff")
        self.style.configure("TButton", background="#00a67d", foreground="#ffffff",
                             relief="flat", borderwidth=0)
        self.style.map("TButton", background=[("active", "#45A049")])
        self.style.configure("TCheckbutton", background="#2e2e2e", foreground="#ffffff")
        self.style.map("TCheckbutton",
                       background=[("active", "#2e2e2e")],
                       foreground=[("active", "#ffffff")])
        self.style.configure("TEntry", fieldbackground="#3e3e3e", foreground="#ffffff")
        self.style.configure("TCombobox", fieldbackground="#3e3e3e", foreground="#ffffff")
        self.style.map("TCombobox",
                       fieldbackground=[("readonly", "#3e3e3e")],
                       foreground=[("readonly", "#ffffff")])
        self.style.configure("TRadiobutton", background="#2e2e2e", foreground="#ffffff")
        self.style.map("TRadiobutton",
                       background=[("active", "#2e2e2e")],
                       foreground=[("active", "#ffffff")])

        self.style.configure("Vertical.TScrollbar", troughcolor="#2e2e2e", bordercolor="#2e2e2e",
                             background="#2e2e2e", arrowcolor="#ffffff")
        self.style.map("Vertical.TScrollbar",
                       background=[("active", "#45A049")],
                       arrowcolor=[("active", "#ffffff")])

        # Variables
        self.combine_dir = tk.StringVar()
        self.combine_mode = tk.StringVar(value="one_sheet")
        self.select_all_var = tk.BooleanVar(value=True)
        self.files_vars = {}

        self.split_file_path = tk.StringVar()
        self.split_sheet = tk.StringVar()
        self.col_combo = None
        self.split_mode = tk.StringVar(value="files")

        self.notebook = ttk.Notebook(self, style="TNotebook")
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.create_combine_tab()
        self.create_split_tab()

    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width() or 1000
        height = self.winfo_height() or 700
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    # ---------- Combine Tab ---------- #
    def create_combine_tab(self):
        frame = ttk.Frame(self.notebook, style="TFrame")
        self.notebook.add(frame, text="Combine Files")

        # Directory row
        dir_frame = ttk.Frame(frame, style="TFrame")
        dir_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_dir = ttk.Label(dir_frame, text="Directory:")
        lbl_dir.pack(side=tk.LEFT, anchor="w")
        entry_dir = ttk.Entry(dir_frame, textvariable=self.combine_dir, width=50)
        entry_dir.pack(side=tk.LEFT, anchor="w", padx=5)
        btn_browse = ttk.Button(dir_frame, text="Browse", command=self.browse_directory)
        btn_browse.pack(side=tk.LEFT, anchor="w", padx=5)

        # Combine mode row
        mode_frame = ttk.Frame(frame, style="TFrame")
        mode_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_mode = ttk.Label(mode_frame, text="Combine Mode:")
        lbl_mode.pack(side=tk.LEFT, anchor="w")
        rbtn_one = ttk.Radiobutton(mode_frame, text="One Sheet", variable=self.combine_mode, value="one_sheet")
        rbtn_one.pack(side=tk.LEFT, anchor="w", padx=5)
        rbtn_sep = ttk.Radiobutton(mode_frame, text="Separate Sheets", variable=self.combine_mode,
                                   value="separate_sheets")
        rbtn_sep.pack(side=tk.LEFT, anchor="w", padx=5)

        # File list row
        list_frame = ttk.Labelframe(frame, text="Excel Files Found", style="TLabelframe")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        header_frame = ttk.Frame(list_frame, style="TFrame")
        header_frame.pack(fill=tk.X, padx=5, pady=5)
        chk_all = ttk.Checkbutton(header_frame, text="Select/Deselect All",
                                  variable=self.select_all_var, command=self.update_all_checkbuttons)
        chk_all.pack(side=tk.LEFT, anchor="w")

        self.canvas = tk.Canvas(list_frame, bg="#2e2e2e", highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.canvas.yview, style="Vertical.TScrollbar")
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.bind('<Configure>', lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.checklist_frame = ttk.Frame(self.canvas, style="TFrame")
        self.canvas.create_window((0, 0), window=self.checklist_frame, anchor="nw")

        # Combine button
        btn_combine = ttk.Button(frame, text="Combine Files", command=self.combine_files_action)
        btn_combine.pack(anchor="w", padx=10, pady=10)

    def browse_directory(self):
        dir_selected = filedialog.askdirectory()
        if dir_selected:
            self.combine_dir.set(dir_selected)
            self.load_files_list()

    def load_files_list(self):
        directory = self.combine_dir.get()
        for widget in self.checklist_frame.winfo_children():
            widget.destroy()
        self.files_vars.clear()

        if not directory or not os.path.isdir(directory):
            messagebox.showerror("Error", "Please select a valid directory.")
            return
        files = [file for file in glob.glob(os.path.join(directory, "*.xlsx"))
                 if not os.path.basename(file).startswith("~")]
        if not files:
            messagebox.showinfo("No Files", "No Excel (.xlsx) files found in the directory.")
            return

        for file in files:
            var = tk.BooleanVar(value=self.select_all_var.get())
            chk = ttk.Checkbutton(self.checklist_frame, text=os.path.basename(file), variable=var)
            chk.pack(anchor="w")
            self.files_vars[file] = var

    def update_all_checkbuttons(self):
        for var in self.files_vars.values():
            var.set(self.select_all_var.get())

    def combine_files_action(self):
        selected_files = [file for file, var in self.files_vars.items() if var.get()]
        if not selected_files:
            messagebox.showerror("Error", "Please select at least one file from the list.")
            return
        mode = self.combine_mode.get()
        output = combine_files(selected_files, combine_mode=mode)
        if output:
            messagebox.showinfo("Success", f"Files combined successfully.\nOutput File:\n{output}")
        else:
            messagebox.showerror("Error", "An error occurred while combining files.")

    # ---------- Split Tab ---------- #
    def create_split_tab(self):
        frame = ttk.Frame(self.notebook, style="TFrame")
        self.notebook.add(frame, text="Split File")

        # File selection
        file_frame = ttk.Frame(frame, style="TFrame")
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_file = ttk.Label(file_frame, text="Excel File:")
        lbl_file.pack(side=tk.LEFT, anchor="w")
        entry_file = ttk.Entry(file_frame, textvariable=self.split_file_path, width=50)
        entry_file.pack(side=tk.LEFT, anchor="w", padx=5)
        btn_browse = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        btn_browse.pack(side=tk.LEFT, anchor="w", padx=5)

        # Sheet selection
        sheet_frame = ttk.Frame(frame, style="TFrame")
        sheet_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_sheet = ttk.Label(sheet_frame, text="Sheet:")
        lbl_sheet.pack(side=tk.LEFT, anchor="w")
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.split_sheet, state="readonly")
        self.sheet_combo.pack(side=tk.LEFT, anchor="w", padx=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        # Column selection
        col_frame = ttk.Frame(frame, style="TFrame")
        col_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_col = ttk.Label(col_frame, text="Split by Column:")
        lbl_col.pack(side=tk.LEFT, anchor="w")
        self.col_combo = ttk.Combobox(col_frame, state="readonly")
        self.col_combo.pack(side=tk.LEFT, anchor="w", padx=5)

        # Split mode
        mode_frame = ttk.Frame(frame, style="TFrame")
        mode_frame.pack(fill=tk.X, padx=10, pady=5)
        lbl_mode = ttk.Label(mode_frame, text="Split Mode:")
        lbl_mode.pack(side=tk.LEFT, anchor="w")
        rbtn_files = ttk.Radiobutton(mode_frame, text="Files", variable=self.split_mode, value="files")
        rbtn_files.pack(side=tk.LEFT, anchor="w", padx=5)
        rbtn_sheets = ttk.Radiobutton(mode_frame, text="Sheets", variable=self.split_mode, value="sheets")
        rbtn_sheets.pack(side=tk.LEFT, anchor="w", padx=5)

        # Split button
        btn_split = ttk.Button(frame, text="Split File", command=self.split_file_action)
        btn_split.pack(anchor="w", padx=10, pady=10)

    def browse_file(self):
        file_selected = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_selected:
            self.split_file_path.set(file_selected)
            self.load_sheet_options_and_columns()

    def load_sheet_options_and_columns(self):
        file_path = self.split_file_path.get()
        if not file_path or not os.path.isfile(file_path):
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return
        sheets = get_sheets_from_file(file_path)
        if not sheets:
            messagebox.showerror("Error", "Could not load sheets from the file.")
            return
        self.sheet_combo["values"] = sheets
        self.sheet_combo.current(0)
        self.split_sheet.set(sheets[0])
        self.load_columns()

    def load_columns(self):
        file_path = self.split_file_path.get()
        sheet = self.split_sheet.get()
        if not file_path or not os.path.isfile(file_path):
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return
        df = read_sheet_from_file(file_path, sheet)
        if df is None:
            messagebox.showerror("Error", "Could not load columns from the file.")
            return
        columns = list(df.columns)
        if not columns:
            messagebox.showerror("Error", "No columns found in the selected sheet.")
            return
        self.col_combo["values"] = columns
        self.col_combo.current(0)

    def on_sheet_change(self, event):
        self.load_columns()

    def split_file_action(self):
        file_path = self.split_file_path.get()
        sheet = self.split_sheet.get()
        selected_column = self.col_combo.get()
        if not file_path or not sheet or not selected_column:
            messagebox.showerror("Error", "Please select a file, sheet, and column.")
            return
        mode = self.split_mode.get()
        output = split_file(file_path, sheet, selected_column, split_mode=mode)
        if output:
            if mode == "files":
                if isinstance(output, list):
                    message = "Data split into files successfully.\nFiles:\n" + "\n".join(output)
                else:
                    message = "Data split into files successfully.\nFile:\n" + str(output)
            else:
                message = f"Data split into sheets successfully.\nOutput File:\n{output}"
            messagebox.showinfo("Success", message)
        else:
            messagebox.showerror("Error", "An error occurred while splitting the file.")


def main():
    app = ExcelManagerGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
