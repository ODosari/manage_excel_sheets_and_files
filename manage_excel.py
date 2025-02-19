#!/usr/bin/env python3
"""
Manage Excel Sheets and Files Utility - CLI

This script allows you to combine multiple Excel files into one,
or split a single Excel file into multiple sheets or files based on a specific column.
It includes enhanced handling for password-protected files.

Version = "0.2"
By = "Obaid Aldosari"
GitHub: https://github.com/ODosari/manage_excel_sheets_and_files
"""

import os
import sys
import time
import glob
import logging
import argparse
import tempfile
import io
import traceback
from contextlib import closing

import pandas as pd
import msoffcrypto
from openpyxl import load_workbook
from msoffcrypto.exceptions import InvalidKeyError

# Configure logging with two handlers:
# - A file handler to log detailed INFO messages.
# - A console handler to show only warnings and errors.
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# File handler: logs INFO and above to 'manage_excel.log'
file_handler = logging.FileHandler("manage_excel.log")
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Console handler: logs only WARNING and above to stderr
console_handler = logging.StreamHandler(sys.stderr)
console_handler.setLevel(logging.WARNING)
console_formatter = logging.Formatter("%(levelname)s: %(message)s")
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)


def print_help_message():
    """
    Prints the help message explaining the utility and its commands.
    """
    help_message = """
################################################################################
Welcome to Manage Excel Sheets and Files Utility!

This utility allows you to combine multiple Excel files into one,
or split a single Excel file into multiple sheets or files based on a specific column.

Commands:
  C <path> - Combine all Excel files in <path> into a single file.
  S <file> - Split an Excel file into multiple sheets or files based on a column.
  Q      - Quit the program.
################################################################################
"""
    print(help_message)


def print_commands():
    """
    Prints the available commands.
    """
    commands = """
Available commands:
  C <path> - Combine all Excel files in <path> into a single file.
  S <file> - Split an Excel file into multiple sheets or files based on a column.
  Q      - Quit the program.
"""
    print(commands)


def get_timestamped_filename(base_path, prefix, extension):
    """
    Generate a unique filename using a timestamp.
    """
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    return os.path.join(base_path, f'{prefix}_{timestamp}{extension}')


def normalize_sheet_name(base_name, sheet_name, existing_names):
    """
    Generate a safe Excel sheet name by combining the base file name and sheet name,
    replacing spaces with underscores and truncating to 31 characters. If the generated
    name already exists, appends a counter.
    """
    raw_name = f"{base_name}_{sheet_name}".replace(" ", "_")
    safe_name = raw_name[:31]
    counter = 1
    while safe_name in existing_names:
        suffix = f"_{counter}"
        safe_name = (raw_name[:31 - len(suffix)] + suffix)
        counter += 1
    existing_names.add(safe_name)
    return safe_name


def unprotect_excel_file(file_path, default_password=None, max_attempts=3):
    """
    Unprotects an Excel file (if encrypted or protected) and returns a path
    to a temporary unprotected copy.
    Supports using a default password and up to max_attempts attempts.
    """
    try:
        with open(file_path, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            if office_file.is_encrypted():
                attempts = 0
                password = default_password
                while attempts < max_attempts:
                    try:
                        if password is None:
                            password = input(f"Enter password for {os.path.basename(file_path)}: ")
                        decrypted = io.BytesIO()
                        office_file.load_key(password=password)
                        office_file.decrypt(decrypted)
                        decrypted.seek(0)
                        wb = load_workbook(decrypted, read_only=False, keep_vba=True)
                        logger.info(f"Password accepted for {os.path.basename(file_path)}")
                        break
                    except InvalidKeyError as ike:
                        attempts += 1
                        logger.error(f"Incorrect password for {os.path.basename(file_path)} (Attempt {attempts}/{max_attempts}): {ike}")
                        if attempts < max_attempts:
                            password = None
                        else:
                            raise Exception("Maximum password attempts reached.") from ike
            else:
                wb = load_workbook(file_path, read_only=False, keep_vba=True)

        wb.security = None
        for sheet in wb.worksheets:
            sheet.protection.enabled = False
            sheet.protection.sheet = False

        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        logger.debug(f"Unprotected file saved to temporary file: {temp_file.name}")
        return temp_file.name

    except Exception as e:
        logger.error(f"Failed to open {os.path.basename(file_path)}: {e}")
        traceback.print_exc()
        return None


def choose_sheet_from_file(file_path, default_password=None):
    """
    Allows the user to choose which sheet(s) to work with from an Excel file.
    """
    temp_file = None
    try:
        temp_file = unprotect_excel_file(file_path, default_password=default_password)
        if temp_file is None:
            return None

        while True:
            with closing(pd.ExcelFile(temp_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                print(f"\nAvailable sheets in {os.path.basename(file_path)}:")
                for idx, sheet in enumerate(sheet_names, start=1):
                    print(f"  {idx}. {sheet}")

                if len(sheet_names) == 1:
                    logger.info(f"Only one sheet available ('{sheet_names[0]}'), automatically selecting it.")
                    df = pd.read_excel(temp_file, sheet_name=sheet_names[0], engine='openpyxl')
                    return [(sheet_names[0], df)]
                else:
                    print("\nType 'all' to select all sheets or 'Q' to skip this file.")
                    user_choice = input(f"Enter the index numbers of the sheets to select from {os.path.basename(file_path)} (separated by commas), 'all', or 'Q' to skip: ").strip()
                    if user_choice.lower() == 'q':
                        logger.info(f"Skipping file {os.path.basename(file_path)}.")
                        return None
                    elif user_choice.lower() == 'all':
                        return [(sheet, pd.read_excel(temp_file, sheet_name=sheet, engine='openpyxl')) for sheet in sheet_names]
                    else:
                        indices = user_choice.split(',')
                        selected = []
                        for idx in indices:
                            idx = idx.strip()
                            if idx.isdigit():
                                index = int(idx)
                                if 1 <= index <= len(sheet_names):
                                    sheet_name = sheet_names[index - 1]
                                    df = pd.read_excel(temp_file, sheet_name=sheet_name, engine='openpyxl')
                                    selected.append((sheet_name, df))
                                else:
                                    print(f"Invalid index number: {index}.")
                            else:
                                print(f"Invalid input: {idx}. Please enter numbers only.")
                        if selected:
                            return selected
                        else:
                            print("No valid sheets selected. Please try again or type 'Q' to skip this file.")
    except Exception as e:
        logger.error(f"Error reading file {os.path.basename(file_path)}: {e}")
        traceback.print_exc()
        return None
    finally:
        if temp_file and os.path.exists(temp_file):
            os.unlink(temp_file)


def combine_excel_files(directory_path, default_password=None):
    """
    Combines selected Excel files from the specified directory.
    """
    try:
        logger.info("Searching for Excel files...")
        files = glob.glob(os.path.join(directory_path, '*.xlsx'))
        if not files:
            logger.warning("No Excel files found in the directory.")
            return

        logger.info("Found the following Excel files:")
        for i, file_path in enumerate(files, 1):
            logger.info(f"  {i}. {os.path.basename(file_path)}")
        print("Type 'all' to select all files.")

        while True:
            selected_input = input("Enter the numbers of the files to combine (separated by commas) or type 'all': ").strip()
            if selected_input.lower() == 'q':
                logger.info("Operation cancelled by the user.")
                return
            if selected_input.lower() == 'all':
                selected_files = files
                break
            else:
                indices = [i.strip() for i in selected_input.split(',')]
                if all(idx.isdigit() and 1 <= int(idx) <= len(files) for idx in indices):
                    selected_files = [files[int(i) - 1] for i in indices]
                    break
                else:
                    print("Invalid input. Please enter valid file numbers or 'all'. (Type 'Q' to cancel)")

        while True:
            choice = input("Combine into one sheet (O) or into one workbook with different sheets (W)? (Type 'Q' to cancel): ").lower()
            if choice == 'q':
                logger.info("Operation cancelled by the user.")
                return
            elif choice in ['o', 'w']:
                break
            else:
                print("Invalid choice. Please enter 'O', 'W', or 'Q' to cancel.")

        output_file = get_timestamped_filename(directory_path, 'Combined', '.xlsx')
        base_names_used = set()

        if choice == 'o':
            combined_df = pd.DataFrame()
            for file_path in selected_files:
                df_list = choose_sheet_from_file(file_path, default_password=default_password) or []
                for sheet_name, df in df_list:
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
            combined_df.to_excel(output_file, sheet_name='Combined', engine='openpyxl', index=False)
        elif choice == 'w':
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for file_path in selected_files:
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    df_list = choose_sheet_from_file(file_path, default_password=default_password) or []
                    for sheet_name, df in df_list:
                        safe_sheet = normalize_sheet_name(base_name, sheet_name, base_names_used)
                        df.to_excel(writer, sheet_name=safe_sheet, index=False)

        logger.info(f"Files combined successfully. Output File: {os.path.basename(output_file)}")

    except Exception as e:
        logger.error(f"An error occurred while combining files: {e}")
        traceback.print_exc()


def split_excel_file(file_path, default_password=None):
    """
    Splits an Excel file into multiple files or sheets based on a column's unique values.
    """
    temp_file = None
    try:
        temp_file = unprotect_excel_file(file_path, default_password=default_password)
        if temp_file is None:
            return

        while True:
            with closing(pd.ExcelFile(temp_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                print(f"\nAvailable sheets in {os.path.basename(file_path)}:")
                for idx, sheet in enumerate(sheet_names, start=1):
                    print(f"  {idx}. {sheet}")

                if len(sheet_names) == 1:
                    chosen_sheet = sheet_names[0]
                    logger.info(f"Only one sheet ('{chosen_sheet}') available, automatically selecting it.")
                else:
                    print("Type 'Q' to skip this file.")
                    chosen_input = input(f"Enter the index number of the sheet to split from {os.path.basename(file_path)}: ").strip()
                    if chosen_input.lower() == 'q':
                        logger.info(f"Skipping file {os.path.basename(file_path)}.")
                        return
                    if not chosen_input.isdigit():
                        print("Invalid input. Please enter a valid index number or 'Q' to skip.")
                        continue
                    index = int(chosen_input)
                    if 1 <= index <= len(sheet_names):
                        chosen_sheet = sheet_names[index - 1]
                    else:
                        print("Invalid index number. Please try again or type 'Q' to skip.")
                        continue

                df = pd.read_excel(temp_file, sheet_name=chosen_sheet, engine='openpyxl')
                columns = df.columns.tolist()

                print("\nColumns available for splitting:")
                for index, col in enumerate(columns, 1):
                    print(f"  {index}. {col}")

                while True:
                    col_input = input("Enter the index number of the column to split by (or type 'Q' to skip): ").strip()
                    if col_input.lower() == 'q':
                        logger.info("Skipping splitting operation.")
                        return
                    if not col_input.isdigit():
                        print("Invalid input. Please enter a number or 'Q' to skip.")
                        continue
                    col_index = int(col_input)
                    if 1 <= col_index <= len(columns):
                        break
                    else:
                        print("Invalid column index. Please try again or type 'Q' to skip.")

                split_column = columns[col_index - 1]
                unique_values = df[split_column].unique()
                logger.info(f"Data will be split based on the values in column '{split_column}': {', '.join(map(str, unique_values))}")

                while True:
                    split_type = input("Split into different Sheets or Files (S/F)? (Type 'Q' to skip): ").lower()
                    if split_type == 'q':
                        logger.info("Skipping splitting operation.")
                        return
                    elif split_type == 'f':
                        send_to_file(df, unique_values, split_column, file_path, chosen_sheet)
                        break
                    elif split_type == 's':
                        send_to_sheet(df, unique_values, split_column, file_path, chosen_sheet)
                        break
                    else:
                        print("Invalid choice. Please enter 'S', 'F', or 'Q' to skip.")
                break

    except Exception as e:
        logger.error(f"An error occurred while splitting the file: {e}")
        traceback.print_exc()
    finally:
        if temp_file and os.path.exists(temp_file):
            os.unlink(temp_file)


def send_to_file(df, unique_values, split_column, original_file, sheet_name):
    """
    Splits the DataFrame into separate files based on unique values in a column.
    """
    directory = os.path.dirname(original_file)
    base_filename = f"{os.path.splitext(os.path.basename(original_file))[0]}_{sheet_name}"
    os.makedirs(directory, exist_ok=True)

    for value in unique_values:
        safe_value = str(value).replace(" ", "_")
        output_file = get_timestamped_filename(directory, f'{base_filename}_{split_column}_{safe_value}', '.xlsx')
        filtered_df = df[df[split_column] == value]
        filtered_df.to_excel(output_file, sheet_name=str(value)[:31], engine='openpyxl', index=False)
        logger.info(f"Data for '{value}' written to {os.path.basename(output_file)}")
    logger.info('Data split into files successfully.')


def send_to_sheet(df, unique_values, split_column, original_file, sheet_name):
    """
    Splits the DataFrame into separate sheets in one workbook based on unique values in a column.
    """
    directory = os.path.dirname(original_file)
    output_file = get_timestamped_filename(directory, f'{os.path.splitext(os.path.basename(original_file))[0]}_{sheet_name}_split', '.xlsx')

    existing_sheet_names = set()
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for value in unique_values:
            raw_sheet_name = str(value)
            safe_sheet = normalize_sheet_name(raw_sheet_name, "", existing_sheet_names) if raw_sheet_name.strip() else "Empty"
            filtered_df = df[df[split_column] == value]
            filtered_df.to_excel(writer, sheet_name=safe_sheet, index=False)
    logger.info(f"Data split into sheets successfully. Output File: {os.path.basename(output_file)}")


def interactive_mode(default_password=None):
    """
    Provides an interactive CLI mode for the user.
    """
    print_help_message()
    while True:
        print_commands()
        user_input = input("Enter your command: ").strip()
        if user_input.lower() == 'q':
            logger.info("Exiting the program.")
            break
        elif user_input.lower() == 'help':
            print_help_message()
            continue
        elif user_input.lower().startswith(('c ', 's ')):
            parts = user_input.strip().split(maxsplit=1)
            if len(parts) < 2:
                print("Please provide a path or file after the command.")
                continue
            operation, path = parts
            if operation.lower() == 'c':
                combine_excel_files(path, default_password=default_password)
            elif operation.lower() == 's':
                split_excel_file(path, default_password=default_password)
            else:
                print("Invalid command. Type 'Q' to quit or 'help' for instructions.")
        else:
            print("Invalid command. Type 'Q' to quit or 'help' for instructions.")


def parse_arguments():
    """
    Parses command-line arguments.
    """
    parser = argparse.ArgumentParser(description="Manage Excel Sheets and Files")
    parser.add_argument('-c', '--combine', help='Path to combine Excel files')
    parser.add_argument('-s', '--split', help='File to split into multiple sheets or files')
    parser.add_argument('-p', '--password', help='Default password for password-protected files', default=None)
    return parser.parse_args()


def main():
    """
    Main entry point of the application.
    """
    args = parse_arguments()
    if args.combine:
        combine_excel_files(args.combine, default_password=args.password)
    elif args.split:
        split_excel_file(args.split, default_password=args.password)
    else:
        interactive_mode(default_password=args.password)


if __name__ == "__main__":
    main()
