#!/usr/bin/env python3

# Name = "Manage Excel Sheets and Files"
# Version = "0.2"
# By = "Obaid Aldosari"
# GitHub = "https://github.com/ODosari/manage_excel_sheets_and_files"

import os
import re
import time
import glob
import argparse
import tempfile
import traceback
import io

import pandas as pd
from contextlib import closing
from openpyxl import load_workbook
import msoffcrypto

def print_help_message():
    print("################################################################################")
    print("Welcome to Manage Excel Sheets and Files Utility!")
    print("This utility allows you to:")
    print(" - Combine multiple Excel files into one (match columns by header).")
    print(" - Split a single Excel file into multiple sheets or files based on a specific column.")
    print("\nCommands:")
    print("  C <path> - Combine all Excel files in <path> into a single file.")
    print("  S <file> - Split an Excel file into multiple sheets or files based on a column.")
    print("  Q       - Quit the program.")
    print("################################################################################")

def print_commands():
    print("\nAvailable commands:")
    print("  C <path> - Combine all Excel files in <path> into a single file.")
    print("  S <file> - Split an Excel file into multiple sheets or files based on a column.")
    print("  Q        - Quit the program.")

def get_timestamped_filename(base_path, prefix, extension):
    """
    Generate a filename with a time-based stamp down to seconds, ensuring we
    don't overwrite existing files. For large loops, add a small uniqueness factor if needed.
    """
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    return os.path.join(base_path, f"{prefix}_{timestamp}{extension}")

def sanitize_filename_part(value):
    """
    Replaces or removes characters that are invalid in Windows, macOS, or Linux filenames.
    """
    # This regex removes any of the following characters: \ / * ? : " < > |
    # Also strip leading/trailing spaces for cleanliness
    safe_str = re.sub(r'[\\/*?:"<>|]', '_', str(value).strip())
    # Optionally you can shorten it if you worry about OS path length limits
    return safe_str[:100]  # trim to 100 chars if you want

def unprotect_excel_file(file):
    """
    Attempts to remove workbook and sheet protection, prompting for a password if needed.
    Returns a path to a temporary unprotected .xlsx copy or None if it fails.
    """
    try:
        with open(file, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            if office_file.is_encrypted():
                password = input(f"Enter password for {os.path.basename(file)}: ")
                decrypted = io.BytesIO()
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                wb = load_workbook(decrypted, read_only=False, keep_vba=True, data_only=False, keep_links=False)
            else:
                # File is not encrypted
                wb = load_workbook(file, read_only=False, keep_vba=True, data_only=False, keep_links=False)

        # Unprotect workbook
        wb.security = None  # Remove workbook-level protection

        # Unprotect sheets
        for sheet in wb.worksheets:
            sheet.protection.enabled = False
            sheet.protection.sheet = False

        # Save to a temporary file
        temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp.name)
        temp.close()
        return temp.name
    except Exception as e:
        print(f"Failed to open {os.path.basename(file)}: {e}")
        return None

def choose_sheet_from_file(file):
    """
    Prompts the user to select one or more sheets from the given Excel file.
    Returns a list of (sheet_name, DataFrame) or None if skipped or error.
    """
    unprotected_file = None
    try:
        unprotected_file = unprotect_excel_file(file)
        if unprotected_file is None:
            return None

        while True:
            with closing(pd.ExcelFile(unprotected_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                print(f"\nAvailable sheets in {os.path.basename(file)}:")
                for idx, sheet_name in enumerate(sheet_names, start=1):
                    print(f"{idx}. {sheet_name}")

                # If only one sheet, just return that sheet automatically
                if len(sheet_names) == 1:
                    print(f"Only one sheet available ('{sheet_names[0]}'), automatically selecting it.")
                    df = pd.read_excel(unprotected_file, sheet_name=sheet_names[0], engine='openpyxl')
                    return [(sheet_names[0], df)]
                else:
                    print("\nType 'all' to select all sheets, or 'Q' to skip this file.")
                    chosen_input = input(
                        f"Enter the sheet numbers to select from {os.path.basename(file)} "
                        "(separated by commas), 'all', or 'Q': "
                    ).strip()

                    if chosen_input.lower() == 'q':
                        print(f"Skipping file {os.path.basename(file)}.")
                        return None
                    elif chosen_input.lower() == 'all':
                        df_list = []
                        for sheet in sheet_names:
                            df = pd.read_excel(unprotected_file, sheet_name=sheet, engine='openpyxl')
                            df_list.append((sheet, df))
                        return df_list
                    else:
                        indices = [x.strip() for x in chosen_input.split(',')]
                        df_list = []
                        invalid_found = False
                        for idx_s in indices:
                            if idx_s.isdigit():
                                index_val = int(idx_s)
                                if 1 <= index_val <= len(sheet_names):
                                    chosen_sheet = sheet_names[index_val - 1]
                                    df = pd.read_excel(unprotected_file, sheet_name=chosen_sheet, engine='openpyxl')
                                    df_list.append((chosen_sheet, df))
                                else:
                                    print(f"Invalid index number {index_val}.")
                                    invalid_found = True
                            else:
                                print(f"Invalid input '{idx_s}'. Please enter numbers only.")
                                invalid_found = True

                        if df_list and not invalid_found:
                            return df_list
                        else:
                            # Let user retry or skip
                            print("No valid sheets selected or some invalid selections. Please try again.")
                            continue
    except Exception as e:
        print(f"Error reading file {os.path.basename(file)}: {e}")
        traceback.print_exc()
        return None
    finally:
        if unprotected_file:
            os.unlink(unprotected_file)

def better_combine_dataframes_by_header(dataframes):
    """
    Takes a list of DataFrames and concatenates them by matching their column headers.
    Missing columns are filled with NaN.
    Returns a single DataFrame or None if the list is empty.
    """
    if not dataframes:
        return None

    # 1) Determine the union of all columns
    all_columns = set()
    for df in dataframes:
        all_columns |= set(df.columns)
    all_columns = list(all_columns)  # Convert set to list for consistent order

    # 2) Reindex each DataFrame to align with the union of columns
    aligned_dfs = []
    for df in dataframes:
        # If you want to enforce intersection only, do:
        #   df = df.reindex(columns=common_cols)
        # but here we do union -> reindex with all_columns
        aligned_dfs.append(df.reindex(columns=all_columns))

    # 3) Concatenate them in one shot
    combined_df = pd.concat(aligned_dfs, ignore_index=True)
    return combined_df

def combine_excel_files(path):
    """
    Combines multiple Excel files into either:
      - One sheet (by union of columns, all rows appended), or
      - One workbook with multiple sheets (no column alignment needed across sheets).
    """
    try:
        print("Searching for Excel files...")
        files = glob.glob(os.path.join(path, '*.xlsx'))
        if not files:
            print("No Excel files found in the directory.")
            return

        print("Found the following Excel files:")
        for i, file in enumerate(files, 1):
            print(f"{i}. {os.path.basename(file)}")
        print("Type 'all' to select all files, or 'Q' to cancel.")

        while True:
            selected_files_idx = input("Enter the numbers of the files to combine (e.g. '1,2') or 'all': ")
            if selected_files_idx.lower() == 'q':
                print("Operation cancelled by the user.")
                return
            if selected_files_idx.lower() == 'all':
                selected_files = files
                break
            else:
                indices = [x.strip() for x in selected_files_idx.split(',')]
                if all(idx.isdigit() and 1 <= int(idx) <= len(files) for idx in indices):
                    selected_files = [files[int(i) - 1] for i in indices]
                    break
                else:
                    print("Invalid input. Please enter valid file numbers or 'all'. Type 'Q' to cancel.")

        if not selected_files:
            print("No files selected. Aborting.")
            return

        # Ask user if they want all data in one single sheet or multiple sheets
        while True:
            choice = input("Combine into one sheet (O) or into one workbook with different sheets (W)? (Q to cancel): ").lower()
            if choice == 'q':
                print("Operation cancelled by the user.")
                return
            elif choice in ['o', 'w']:
                break
            else:
                print("Invalid choice. Please enter 'O', 'W', or 'Q' to cancel.")

        output_file = get_timestamped_filename(path, 'Combined', '.xlsx')

        if choice == 'o':
            # Combine all chosen files/sheets into ONE sheet
            all_dfs = []
            for file in selected_files:
                df_list = choose_sheet_from_file(file)
                if df_list is not None:
                    # df_list is a list of (sheet_name, df)
                    for (_, df) in df_list:
                        all_dfs.append(df)

            if not all_dfs:
                print("No DataFrames were selected or found. Nothing to combine.")
                return

            # Combine by matching columns (union)
            combined_df = better_combine_dataframes_by_header(all_dfs)
            if combined_df is not None:
                combined_df.to_excel(output_file, sheet_name='Combined', engine='openpyxl', index=False)
                print(f'Files combined successfully into one sheet.\nOutput File: {os.path.basename(output_file)}')
            else:
                print("No data was combined (perhaps no valid sheets).")

        elif choice == 'w':
            # Combine all chosen files, each selected sheet -> separate sheet in one workbook
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for file in selected_files:
                    df_list = choose_sheet_from_file(file)
                    if df_list is not None:
                        for sheet_name, df in df_list:
                            # Create a safe sheet name (max 31 chars, no invalid chars)
                            base_part = sanitize_filename_part(os.path.splitext(os.path.basename(file))[0])
                            sheet_part = sanitize_filename_part(sheet_name)
                            safe_sheet_name = (f"{base_part}_{sheet_part}")[:31]
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            print(f'Files combined successfully into one workbook (multiple sheets).\nOutput File: {os.path.basename(output_file)}')

    except Exception as e:
        print(f"An error occurred while combining files: {e}")
        traceback.print_exc()

def split_excel_file(file):
    """
    Splits an Excel file into multiple sheets or files based on a column the user chooses.
    """
    unprotected_file = None
    try:
        unprotected_file = unprotect_excel_file(file)
        if unprotected_file is None:
            return

        while True:
            with closing(pd.ExcelFile(unprotected_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                print(f"\nAvailable sheets in {os.path.basename(file)}:")
                for idx, sheet_name in enumerate(sheet_names, start=1):
                    print(f"{idx}. {sheet_name}")

                if len(sheet_names) == 1:
                    chosen_sheet = sheet_names[0]
                    print(f"Only one sheet ('{chosen_sheet}') available, automatically selected.")
                else:
                    print("Type 'Q' to skip this file.")
                    chosen_input = input(f"Enter the index number of the sheet to split from {os.path.basename(file)}: ").strip()
                    if chosen_input.lower() == 'q':
                        print(f"Skipping file {os.path.basename(file)}.")
                        return
                    if not chosen_input.isdigit():
                        print("Invalid input. Please enter a valid index number or 'Q' to skip.")
                        continue
                    index_val = int(chosen_input)
                    if 1 <= index_val <= len(sheet_names):
                        chosen_sheet = sheet_names[index_val - 1]
                    else:
                        print("Invalid index number. Please try again or type 'Q' to skip.")
                        continue

                # Now read that chosen sheet
                df = pd.read_excel(unprotected_file, sheet_name=chosen_sheet, engine='openpyxl')
                if df.empty:
                    print(f"The chosen sheet '{chosen_sheet}' is empty. Nothing to split.")
                    return

                cols_name = df.columns.tolist()
                print("\nColumns available for splitting:")
                for idx_c, col in enumerate(cols_name, 1):
                    print(f"{idx_c}. {col}")

                while True:
                    column_input = input("Enter the index number of the column to split by (or type 'Q' to skip): ").strip()
                    if column_input.lower() == 'q':
                        print("Skipping splitting operation.")
                        return
                    if not column_input.isdigit():
                        print("Invalid input. Please enter a number or 'Q' to skip.")
                        continue
                    column_index = int(column_input)
                    if 1 <= column_index <= len(cols_name):
                        break
                    else:
                        print("Invalid column index. Please try again or type 'Q' to skip.")

                column_name = cols_name[column_index - 1]
                unique_vals = df[column_name].unique()
                print(f'Data will be split based on unique values in "{column_name}": {", ".join(map(str, unique_vals))}.')

                while True:
                    split_type = input("Split into different Sheets or Files (S/F)? (Type 'Q' to skip): ").lower()
                    if split_type == 'q':
                        print("Skipping splitting operation.")
                        return
                    elif split_type == 'f':
                        send_to_file(df, unique_vals, column_name, file, chosen_sheet)
                        break
                    elif split_type == 's':
                        send_to_sheet(df, unique_vals, column_name, file, chosen_sheet)
                        break
                    else:
                        print("Invalid choice. Please enter 'S', 'F', or 'Q' to skip.")

                # After successful splitting, break out of the while loop
                break

    except Exception as e:
        print(f"An error occurred while splitting the file: {e}")
        traceback.print_exc()
    finally:
        if unprotected_file:
            os.unlink(unprotected_file)

def send_to_file(df, unique_vals, column_name, file, sheet_name):
    """
    Splits a DataFrame into multiple *files* based on unique column values.
    Each file gets the rows belonging to that unique value.
    """
    directory = os.path.dirname(file)
    base_filename = sanitize_filename_part(os.path.splitext(os.path.basename(file))[0])
    base_sheet = sanitize_filename_part(sheet_name)
    os.makedirs(directory, exist_ok=True)

    for val in unique_vals:
        safe_val = sanitize_filename_part(val)
        # A prefix that includes the base filename, sheet name, column, and value
        prefix = f"{base_filename}_{base_sheet}_{sanitize_filename_part(column_name)}_{safe_val}"
        output_file = get_timestamped_filename(directory, prefix, '.xlsx')
        filtered_df = df[df[column_name] == val]
        filtered_df.to_excel(output_file, sheet_name=str(safe_val)[:31], engine='openpyxl', index=False)

    print('Data split into multiple files successfully.')

def send_to_sheet(df, unique_vals, column_name, file, sheet_name):
    """
    Splits a DataFrame into multiple *sheets* in a single output file,
    each sheet corresponds to one unique column value.
    """
    directory = os.path.dirname(file)
    base_filename = sanitize_filename_part(os.path.splitext(os.path.basename(file))[0])
    base_sheet = sanitize_filename_part(sheet_name)
    prefix = f"{base_filename}_{base_sheet}_split"

    output_file = get_timestamped_filename(directory, prefix, '.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for val in unique_vals:
            safe_val = sanitize_filename_part(val)
            filtered_df = df[df[column_name] == val]
            sheet_title = safe_val[:31] if safe_val else "EmptyValue"
            filtered_df.to_excel(writer, sheet_name=sheet_title, index=False)

    print(f'Data split into sheets successfully. Output File: {os.path.basename(output_file)}')

def interactive_mode():
    print_help_message()
    while True:
        print_commands()
        user_input = input("Enter your command: ").strip()
        if user_input.lower() == 'q':
            break
        elif user_input.lower() == 'help':
            print_help_message()
            continue
        elif user_input.startswith(('C ', 'c ', 'S ', 's ')):
            parts = user_input.strip().split(maxsplit=1)
            if len(parts) < 2:
                print("Please provide a path or file after the command.")
                continue
            operation, path = parts
            operation = operation.lower()
            if operation == 'c':
                combine_excel_files(path)
            elif operation == 's':
                split_excel_file(path)
            else:
                print("Invalid command. Type 'Q' to quit or 'help' for usage.")
        else:
            print("Invalid command. Type 'Q' to quit or 'help' for usage.")

def parse_arguments():
    parser = argparse.ArgumentParser(description="Manage Excel Sheets and Files")
    parser.add_argument('-c', '--combine', help='Path to combine Excel files')
    parser.add_argument('-s', '--split', help='File to split into multiple sheets or files')
    return parser.parse_args()

def main():
    args = parse_arguments()
    if args.combine:
        combine_excel_files(args.combine)
    elif args.split:
        split_excel_file(args.split)
    else:
        interactive_mode()

if __name__ == "__main__":
    main()
