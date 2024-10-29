#!/usr/bin/env python3

# Name = "Manage Excel Sheets and Files"
# Version = "1.6"
# By = "Obaid Aldosari"
# GitHub = "https://github.com/ODosari/manage_excel_sheets_and_files"

import os
import pandas as pd
import time
import glob
from openpyxl import load_workbook
import tempfile
import msoffcrypto
import io
from contextlib import closing
import logging

# Configure logging
logging.basicConfig(filename='excel_manager.log', level=logging.ERROR,
                    format='%(asctime)s:%(levelname)s:%(message)s')


def print_welcome_message():
    print("################################################################################")
    print("Welcome to the Manage Excel Sheets and Files Utility!")
    print("This utility allows you to combine multiple Excel/CSV files into one,")
    print("or split a single Excel/CSV file into multiple sheets or files based on a specific column.")
    print("################################################################################")


def print_main_menu():
    print("\nMain Menu:")
    print("1. Combine Excel/CSV files")
    print("2. Split an Excel/CSV file")
    print("3. Help")
    print("4. Quit")


def print_help_message():
    print("\nHelp:")
    print("This utility supports the following operations:")
    print("1. Combine multiple Excel/CSV files into one file.")
    print("   - Files can be combined into one sheet or into a workbook with multiple sheets.")
    print("2. Split a single Excel/CSV file into multiple sheets or files based on a column.")
    print("\nSupported file formats: .xlsx, .xls, .csv")
    print("You can specify an output directory for the resulting files.")
    print("Password-protected Excel files are supported. You will be prompted for passwords if necessary.")
    print("\nAt any prompt, you can type 'Q' to cancel the current operation and return to the main menu.\n")


def get_timestamped_filename(base_path, prefix, extension):
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    return os.path.join(base_path, f'{prefix}_{timestamp}{extension}')


def unprotect_excel_file(file, password=None):
    try:
        with open(file, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            if office_file.is_encrypted():
                if not password:
                    password = input(f"Enter password for {os.path.basename(file)}: ")
                    if password.lower() == 'q':
                        print(f"Skipping file {os.path.basename(file)}.")
                        return None
                decrypted = io.BytesIO()
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                wb = load_workbook(decrypted, read_only=False, keep_vba=True, data_only=False, keep_links=False)
            else:
                # File is not encrypted
                wb = load_workbook(file, read_only=False, keep_vba=True, data_only=False, keep_links=False)

        # Unprotect workbook and sheets
        wb.security = None
        for sheet in wb.worksheets:
            sheet.protection.enabled = False
            sheet.protection.sheet = False

        # Save to a temporary file
        temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp.name)
        temp.close()
        return temp.name
    except Exception as e:
        print(f"Failed to open {os.path.basename(file)}: {e}")
        logging.error(f"Failed to unprotect {file}: {e}")
        return None


def combine_excel_files():
    try:
        path = input("Enter the directory path containing Excel/CSV files to combine: ").strip()
        if path.lower() == 'q':
            print("Operation cancelled. Returning to main menu.")
            return
        if not os.path.isdir(path):
            print("Invalid directory path. Please try again.")
            return

        files = glob.glob(os.path.join(path, '*.xlsx')) + glob.glob(os.path.join(path, '*.xls')) + glob.glob(os.path.join(path, '*.csv'))
        if not files:
            print("No Excel or CSV files found in the directory.")
            return

        print("\nFound the following files:")
        for i, file in enumerate(files, 1):
            print(f"{i}. {os.path.basename(file)}")
        print("\nType 'all' to select all files or enter the numbers separated by commas.")

        selected_files = []
        while True:
            selected_files_idx = input("Enter your choice: ").strip()
            if selected_files_idx.lower() == 'q':
                print("Operation cancelled. Returning to main menu.")
                return
            if selected_files_idx.lower() == 'all':
                selected_files = files
                break
            else:
                indices = [i.strip() for i in selected_files_idx.split(',')]
                if all(idx.isdigit() and 1 <= int(idx) <= len(files) for idx in indices):
                    selected_files = [files[int(i) - 1] for i in indices]
                    break
                else:
                    print("Invalid input. Please enter valid file numbers separated by commas, 'all', or 'Q' to cancel.")

        password_dict = {}
        for file in selected_files.copy():
            if file.endswith(('.xlsx', '.xls')):
                if is_file_encrypted(file):
                    print(f"File {os.path.basename(file)} is encrypted.")
                    password = input(f"Enter password for {os.path.basename(file)} (or 'Q' to skip): ").strip()
                    if password.lower() == 'q':
                        print(f"Skipping encrypted file {os.path.basename(file)} due to missing password.")
                        selected_files.remove(file)
                    else:
                        password_dict[file] = password

        while True:
            choice = input("Combine into one sheet (O) or into one workbook with different sheets (W)? [O/W]: ").lower()
            if choice == 'q':
                print("Operation cancelled. Returning to main menu.")
                return
            elif choice in ['o', 'w']:
                break
            else:
                print("Invalid choice. Please enter 'O' or 'W'. Type 'Q' to cancel.")

        output_dir = input("Enter output directory (leave blank for current directory): ").strip()
        if output_dir.lower() == 'q':
            print("Operation cancelled. Returning to main menu.")
            return
        if not output_dir:
            output_dir = path
        elif not os.path.isdir(output_dir):
            print("Invalid output directory. Using current directory.")
            output_dir = path

        output_file = get_timestamped_filename(output_dir, 'Combined', '.xlsx')

        print("Processing files...")
        if choice == 'o':
            combined_df = pd.DataFrame()
            for file in selected_files:
                df_list = choose_sheet_from_file(file, password_dict.get(file))
                if df_list is not None:
                    for sheet_name, df in df_list:
                        combined_df = pd.concat([combined_df, df], ignore_index=True)
            if not combined_df.empty:
                combined_df.to_excel(output_file, sheet_name='Combined', engine='openpyxl', index=False)
                print(f'Files combined successfully into one sheet. Output File: {os.path.basename(output_file)}')
            else:
                print("No data to combine. Operation aborted.")
        elif choice == 'w':
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for file in selected_files:
                    df_list = choose_sheet_from_file(file, password_dict.get(file))
                    if df_list is not None:
                        for sheet_name, df in df_list:
                            safe_sheet_name = f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}".replace(' ', '_')[:31]
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            print(f'Files combined successfully into one workbook with multiple sheets. Output File: {os.path.basename(output_file)}')

    except Exception as e:
        print(f"An error occurred while combining files: {e}")
        logging.error(f"Error in combine_excel_files: {e}")
    finally:
        pass


def is_file_encrypted(file):
    try:
        with open(file, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            return office_file.is_encrypted()
    except Exception:
        return False


def choose_sheet_from_file(file, password=None):
    unprotected_file = None
    try:
        # For .xlsx and .xls files
        if file.endswith(('.xlsx', '.xls')):
            unprotected_file = unprotect_excel_file(file, password)
            if unprotected_file is None:
                return None

            with closing(pd.ExcelFile(unprotected_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                if not sheet_names:
                    print(f"No sheets found in {os.path.basename(file)}.")
                    return None
                if len(sheet_names) == 1:
                    chosen_sheets = [sheet_names[0]]
                    print(f"\nOnly one sheet ('{sheet_names[0]}') available in {os.path.basename(file)}, automatically selected.")
                else:
                    print(f"\nAvailable sheets in {os.path.basename(file)}:")
                    for idx, sheet_name in enumerate(sheet_names, start=1):
                        print(f"{idx}. {sheet_name}")

                    print("\nType 'all' to select all sheets or enter the numbers separated by commas.")
                    while True:
                        chosen_input = input(f"Enter your choice for {os.path.basename(file)}: ").strip()
                        if chosen_input.lower() == 'q':
                            print(f"Skipping file {os.path.basename(file)}.")
                            return None
                        elif chosen_input.lower() == 'all':
                            chosen_sheets = sheet_names
                            break
                        else:
                            indices = chosen_input.split(',')
                            if all(idx.strip().isdigit() and 1 <= int(idx.strip()) <= len(sheet_names) for idx in indices):
                                chosen_sheets = [sheet_names[int(idx.strip()) - 1] for idx in indices]
                                break
                            else:
                                print("Invalid input. Please enter valid sheet numbers separated by commas, 'all', or 'Q' to skip.")

                df_list = []
                for sheet in chosen_sheets:
                    df = pd.read_excel(unprotected_file, sheet_name=sheet, engine='openpyxl')
                    df_list.append((sheet, df))
                return df_list

        elif file.endswith('.csv'):
            df = pd.read_csv(file)
            return [(os.path.basename(file), df)]
        else:
            print(f"Unsupported file format for file {os.path.basename(file)}.")
            return None

    except Exception as e:
        print(f"Error reading file {os.path.basename(file)}: {e}")
        logging.error(f"Error in choose_sheet_from_file for {file}: {e}")
        return None
    finally:
        if unprotected_file and os.path.exists(unprotected_file):
            os.unlink(unprotected_file)


def split_excel_file():
    try:
        file = input("Enter the path of the Excel/CSV file to split: ").strip()
        if file.lower() == 'q':
            print("Operation cancelled. Returning to main menu.")
            return
        if not os.path.isfile(file):
            print("File not found. Please try again.")
            return

        output_dir = input("Enter output directory (leave blank for same directory as input file): ").strip()
        if output_dir.lower() == 'q':
            print("Operation cancelled. Returning to main menu.")
            return
        if not output_dir:
            output_dir = os.path.dirname(file)
        elif not os.path.isdir(output_dir):
            print("Invalid output directory. Using same directory as input file.")
            output_dir = os.path.dirname(file)

        if file.endswith(('.xlsx', '.xls')):
            if is_file_encrypted(file):
                password = input(f"Enter password for {os.path.basename(file)} (or 'Q' to cancel): ").strip()
                if password.lower() == 'q':
                    print("Operation cancelled. Returning to main menu.")
                    return
                unprotected_file = unprotect_excel_file(file, password)
            else:
                unprotected_file = unprotect_excel_file(file)
            if unprotected_file is None:
                return

            workbook_path = unprotected_file
            engine = 'openpyxl'
        elif file.endswith('.csv'):
            workbook_path = file
            engine = None
        else:
            print(f"Unsupported file format: {os.path.basename(file)}.")
            return

        with closing(pd.ExcelFile(workbook_path, engine=engine)) as workbook:
            sheet_names = workbook.sheet_names if engine else [os.path.basename(file)]
            if not sheet_names:
                print(f"No sheets found in {os.path.basename(file)}.")
                return
            if len(sheet_names) == 1:
                chosen_sheet = sheet_names[0]
                print(f"\nOnly one sheet ('{chosen_sheet}') available in {os.path.basename(file)}, automatically selected.")
            else:
                print(f"\nAvailable sheets in {os.path.basename(file)}:")
                for idx, sheet_name in enumerate(sheet_names, start=1):
                    print(f"{idx}. {sheet_name}")
                while True:
                    chosen_input = input(f"Enter the number of the sheet to split from {os.path.basename(file)}: ").strip()
                    if chosen_input.lower() == 'q':
                        print("Operation cancelled. Returning to main menu.")
                        return
                    if chosen_input.isdigit() and 1 <= int(chosen_input) <= len(sheet_names):
                        chosen_sheet = sheet_names[int(chosen_input) - 1]
                        break
                    else:
                        print("Invalid input. Please enter a valid sheet number or 'Q' to cancel.")

        if engine:
            df = pd.read_excel(workbook_path, sheet_name=chosen_sheet, engine=engine)
        else:
            df = pd.read_csv(workbook_path)

        cols_name = df.columns.tolist()
        if not cols_name:
            print("No columns found in the selected sheet.")
            return

        print("\nColumns available for splitting:")
        for index, col in enumerate(cols_name, 1):
            print(f"{index}. {col}")

        while True:
            column_input = input("Enter the number of the column to split by: ").strip()
            if column_input.lower() == 'q':
                print("Operation cancelled. Returning to main menu.")
                return
            if column_input.isdigit() and 1 <= int(column_input) <= len(cols_name):
                column_index = int(column_input)
                break
            else:
                print("Invalid input. Please enter a valid column number or 'Q' to cancel.")

        column_name = cols_name[column_index - 1]
        cols = df[column_name].dropna().unique()
        if cols.size <= 1:
            print(f"Not enough unique values found in column '{column_name}'. Cannot split.")
            return
        print(f'\nYour data will be split based on these values in "{column_name}": {", ".join(map(str, cols))}.')

        while True:
            split_type = input("Split into different Sheets or Files? [S/F]: ").lower()
            if split_type == 'q':
                print("Operation cancelled. Returning to main menu.")
                return
            elif split_type in ['s', 'f']:
                break
            else:
                print("Invalid choice. Please enter 'S' or 'F'. Type 'Q' to cancel.")

        print("Processing split...")
        if split_type == 'f':
            send_to_file(df, cols, column_name, file, chosen_sheet, output_dir)
        elif split_type == 's':
            send_to_sheet(df, cols, column_name, file, chosen_sheet, output_dir)

        print("Splitting completed successfully.")

    except Exception as e:
        print(f"An error occurred while splitting the file: {e}")
        logging.error(f"Error in split_excel_file: {e}")
    finally:
        if 'unprotected_file' in locals() and os.path.exists(unprotected_file):
            os.unlink(unprotected_file)


def send_to_file(df, cols, column_name, file, sheet_name, output_dir):
    try:
        base_filename = f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}"

        os.makedirs(output_dir, exist_ok=True)

        for value in cols:
            output_file = get_timestamped_filename(output_dir, f'{base_filename}_{column_name}_{value}', '.xlsx')
            df_subset = df[df[column_name] == value]
            df_subset.to_excel(output_file, sheet_name=str(value), engine='openpyxl', index=False)
    except Exception as e:
        print(f"An error occurred while saving split files: {e}")
        logging.error(f"Error in send_to_file: {e}")


def send_to_sheet(df, cols, column_name, file, sheet_name, output_dir):
    try:
        output_file = get_timestamped_filename(output_dir, f'{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}_split', '.xlsx')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for value in cols:
                sn = str(value)[:31]  # Excel sheet name limit is 31 characters
                filtered_df = df[df[column_name] == value]
                filtered_df.to_excel(writer, sheet_name=sn, index=False)
    except Exception as e:
        print(f"An error occurred while saving split sheets: {e}")
        logging.error(f"Error in send_to_sheet: {e}")


def main():
    print_welcome_message()
    while True:
        print_main_menu()
        choice = input("Enter your choice: ").strip()
        if choice == '1':
            combine_excel_files()
        elif choice == '2':
            split_excel_file()
        elif choice == '3':
            print_help_message()
        elif choice == '4' or choice.lower() == 'q':
            print("Thank you for using the Manage Excel Sheets and Files Utility. Goodbye!")
            break
        else:
            print("Invalid choice. Please select an option from the menu.")


if __name__ == "__main__":
    main()
