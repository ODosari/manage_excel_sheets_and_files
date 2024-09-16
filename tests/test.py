#!/usr/bin/env python3

# Name = "Manage Excel and CSV Files"
# Version = "0.8"
# By = "Obaid Aldosari"
# GitHub = "https://github.com/ODosari/manage_excel_sheets_and_files"

import sys
import os
import pandas as pd
import time
import glob
from openpyxl import load_workbook
import argparse
import tempfile
import msoffcrypto
import io
import traceback
from contextlib import closing
import logging
import getpass
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed
import re

# Dependency Checks
required_modules = {
    'pandas': 'pd',
    'openpyxl': 'load_workbook',
    'msoffcrypto': 'msoffcrypto',
    'tqdm': 'tqdm'
}

missing_modules = []
for module_name, alias in required_modules.items():
    try:
        __import__(module_name)
    except ImportError:
        missing_modules.append(module_name)

if missing_modules:
    print(f"Missing dependencies: {', '.join(missing_modules)}")
    print("Please install the required packages using 'pip install -r requirements.txt'")
    sys.exit(1)

# Configure logging
logging.basicConfig(
    filename='manage_excel.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Global constants
MAX_PASSWORD_ATTEMPTS = 3
EXCEL_SHEET_NAME_LIMIT = 31

def print_help_message():
    """
    Prints the help message with usage instructions.
    """
    help_message = """
################################################################################
Welcome to Manage Excel and CSV Files Utility!
This utility allows you to combine multiple Excel or CSV files into one,
or split a single file into multiple sheets or files based on a specific column.

Commands:
C <path> [options] - Combine all files in <path> into a single file.
S <file> [options] - Split a file into multiple sheets or files based on a column.
Q - Quit the program.

Options:
--output OUTPUT       Specify the output directory for the results.
--combine-mode MODE   Combine into one sheet (sheet) or one workbook (workbook).
--password PASSWORD   Password for encrypted Excel files.
--sheet-name SHEET    Specify the sheet name to use.
--file-type TYPE      Specify the file type to process (excel, csv, all).

Examples:
C /path/to/directory --combine-mode workbook
S /path/to/file.xlsx --sheet-name Sheet1

For more information, please refer to the documentation.
################################################################################
"""
    print(help_message)

def get_timestamped_filename(base_path, prefix, extension):
    """
    Generates a timestamped filename.

    Parameters:
    base_path (str): The directory path.
    prefix (str): The prefix for the filename.
    extension (str): The file extension.

    Returns:
    str: The generated filename with path.
    """
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    return os.path.join(base_path, f'{prefix}_{timestamp}{extension}')

def sanitize_sheet_name(name):
    """
    Sanitizes the sheet name to comply with Excel limitations.

    Parameters:
    name (str): The original sheet name.

    Returns:
    str: The sanitized sheet name.
    """
    # Remove invalid characters
    name = re.sub(r'[:\\/*?\[\]]', '_', name)
    # Truncate to Excel's sheet name limit
    return name[:EXCEL_SHEET_NAME_LIMIT]

def unprotect_excel_file(file, password=None):
    """
    Unprotects an Excel file by removing workbook and sheet protections.

    Parameters:
    file (str): The path to the Excel file.
    password (str): The password for the encrypted file.

    Returns:
    str: The path to the unprotected temporary file.
    """
    attempts = 0
    while attempts < MAX_PASSWORD_ATTEMPTS:
        try:
            with open(file, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                if office_file.is_encrypted():
                    if password is None:
                        # Use getpass to securely get the password without echoing
                        password = getpass.getpass(f"Enter password for {file}: ")
                    decrypted = io.BytesIO()
                    office_file.load_key(password=password)
                    office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = load_workbook(decrypted, read_only=False, keep_vba=True, data_only=False, keep_links=False)
                else:
                    # File is not encrypted
                    wb = load_workbook(file, read_only=False, keep_vba=True, data_only=False, keep_links=False)
                # Unprotect workbook
                wb.security = None  # Remove workbook-level protection
                # Unprotect sheets
                for sheet in wb.worksheets:
                    sheet.protection.enabled = False
                    sheet.protection.sheet = False
                # Save to a temporary file
                temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                wb.save(temp.name)
                temp.close()
                return temp.name
        except Exception as e:
            attempts += 1
            logging.warning(f"Failed to open {file}: {e}")
            if attempts < MAX_PASSWORD_ATTEMPTS:
                print("Incorrect password. Please try again.")
                password = None  # Reset password for next attempt
            else:
                print(f"Maximum attempts reached for {file}. Skipping file.")
                return None
    return None

def combine_excel_files(path, output_dir=None, combine_mode='workbook', password=None, file_type='excel'):
    """
    Combines multiple Excel or CSV files in a directory into a single file.

    Parameters:
    path (str): The directory path containing files to combine.
    output_dir (str): The output directory for the combined file.
    combine_mode (str): The mode of combining ('sheet' or 'workbook').
    password (str): Password for encrypted Excel files.
    file_type (str): The type of files to process ('excel', 'csv', or 'all').

    Returns:
    None
    """
    try:
        if output_dir is None:
            output_dir = path
        logging.info("Searching for files...")
        # Get list of files based on file_type
        files = []
        if file_type in ('excel', 'all'):
            files.extend(glob.glob(os.path.join(path, '*.xlsx')))
            files.extend(glob.glob(os.path.join(path, '*.xls')))
        if file_type in ('csv', 'all'):
            files.extend(glob.glob(os.path.join(path, '*.csv')))
        if not files:
            print("No files found in the directory.")
            return
        print("Found the following files:")
        for i, file in enumerate(files, 1):
            print(f"{i}. {file}")
        print("Type 'all' to select all files.")
        selected_files_idx = input("Enter the numbers of the files to combine (separated by commas) or type 'all': ")
        selected_files = files if selected_files_idx.lower() == 'all' else [files[int(i) - 1] for i in selected_files_idx.split(',') if i.strip().isdigit() and 0 < int(i) <= len(files)]
        output_file = get_timestamped_filename(output_dir, 'Combined', '.xlsx')
        if combine_mode == 'sheet':
            combined_df = pd.DataFrame()
            for file in tqdm(selected_files, desc="Combining files"):
                df_list = choose_sheet_from_file(file, password)
                if df_list is not None:
                    for sheet_name, df in df_list:
                        combined_df = pd.concat([combined_df, df], ignore_index=True)
            combined_df.to_excel(output_file, sheet_name='Combined', engine='openpyxl', index=False)
        elif combine_mode == 'workbook':
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for file in tqdm(selected_files, desc="Combining files"):
                    df_list = choose_sheet_from_file(file, password)
                    if df_list is not None:
                        for sheet_name, df in df_list:
                            safe_sheet_name = sanitize_sheet_name(f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name}")
                            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        print(f'Files combined successfully. Output File: {output_file}')
        logging.info(f'Files combined successfully. Output File: {output_file}')
    except Exception as e:
        print(f"An error occurred while combining files: {e}")
        logging.error(f"An error occurred while combining files: {e}", exc_info=True)

def choose_sheet_from_file(file, password=None):
    """
    Allows the user to select sheets from a file to process.

    Parameters:
    file (str): The file path.
    password (str): Password for encrypted Excel files.

    Returns:
    list: A list of tuples containing sheet names and DataFrames.
    """
    unprotected_file = None
    try:
        file_ext = os.path.splitext(file)[1].lower()
        if file_ext == '.csv':
            df = pd.read_csv(file)
            return [('Sheet1', df)]
        elif file_ext in ['.xlsx', '.xls']:
            unprotected_file = unprotect_excel_file(file, password)
            if unprotected_file is None:
                return None
            with closing(pd.ExcelFile(unprotected_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                print(f"Available sheets in {file}: {sheet_names}")
                if len(sheet_names) == 1:
                    print(f"Only one sheet available ('{sheet_names[0]}'), automatically selecting it.")
                    df = pd.read_excel(unprotected_file, sheet_name=sheet_names[0], engine='openpyxl')
                    return [(sheet_names[0], df)]
                else:
                    print("Type 'all' to select all sheets.")
                    chosen_sheet = input(f"Enter the name of the sheet to combine from {file} or type 'all': ").strip()
                    if chosen_sheet.lower() == 'all':
                        df_list = []
                        for sheet in sheet_names:
                            df = pd.read_excel(unprotected_file, sheet_name=sheet, engine='openpyxl')
                            df_list.append((sheet, df))
                        return df_list
                    elif chosen_sheet not in sheet_names:
                        print(f"Sheet '{chosen_sheet}' not found in {file}. Skipping this file.")
                        logging.warning(f"Sheet '{chosen_sheet}' not found in {file}.")
                        return None
                    else:
                        df = pd.read_excel(unprotected_file, sheet_name=chosen_sheet, engine='openpyxl')
                        return [(chosen_sheet, df)]
        else:
            print(f"Unsupported file type: {file}")
            logging.warning(f"Unsupported file type: {file}")
            return None
    except Exception as e:
        print(f"Error reading file {file}: {e}")
        logging.error(f"Error reading file {file}: {e}", exc_info=True)
        return None
    finally:
        if unprotected_file:
            os.unlink(unprotected_file)

def split_excel_file(file, output_dir=None, password=None, sheet_name=None, file_type='excel'):
    """
    Splits an Excel or CSV file into multiple sheets or files based on a column.

    Parameters:
    file (str): The file path.
    output_dir (str): The output directory for the split files.
    password (str): Password for encrypted Excel files.
    sheet_name (str): The sheet name to process.
    file_type (str): The type of file ('excel' or 'csv').

    Returns:
    None
    """
    unprotected_file = None
    try:
        if output_dir is None:
            output_dir = os.path.dirname(file)
        file_ext = os.path.splitext(file)[1].lower()
        if file_ext == '.csv' or file_type == 'csv':
            df = pd.read_csv(file, chunksize=10000)
            df = pd.concat(df)  # Concatenate all chunks
            cols_name = df.columns.tolist()
        elif file_ext in ['.xlsx', '.xls']:
            unprotected_file = unprotect_excel_file(file, password)
            if unprotected_file is None:
                return
            with closing(pd.ExcelFile(unprotected_file, engine='openpyxl')) as workbook:
                sheet_names = workbook.sheet_names
                if sheet_name is None:
                    print(f"Available sheets: {sheet_names}")
                    if len(sheet_names) == 1:
                        sheet_name = sheet_names[0]
                        print(f"Only one sheet ('{sheet_name}') available, automatically selected.")
                    else:
                        sheet_name = input("Enter the name of the sheet to split: ").strip()
                        if sheet_name not in sheet_names:
                            print(f"Sheet '{sheet_name}' not found in the workbook. Please try again.")
                            logging.warning(f"Sheet '{sheet_name}' not found in {file}.")
                            return
                df = pd.read_excel(unprotected_file, sheet_name=sheet_name, engine='openpyxl')
                cols_name = df.columns.tolist()
        else:
            print(f"Unsupported file type: {file}")
            logging.warning(f"Unsupported file type: {file}")
            return
        # List columns with indices
        print("Columns available for splitting:")
        for index, col in enumerate(cols_name, 1):
            print(f"{index}. {col}")
        column_index = input('Enter the index number of the column to split by: ')
        if not column_index.isdigit() or not (1 <= int(column_index) <= len(cols_name)):
            print("Invalid column index. Please try again.")
            logging.warning("Invalid column index entered.")
            return
        column_index = int(column_index)
        column_name = cols_name[column_index - 1]
        cols = df[column_name].unique()
        print(f'Your data will be split based on these values in "{column_name}": {", ".join(map(str, cols))}.')
        split_type = input('Split into different Sheets or Files (S/F): ').lower()
        if split_type == 'f':
            send_to_file(df, cols, column_name, file, sheet_name, output_dir)
        elif split_type == 's':
            send_to_sheet(df, cols, column_name, file, sheet_name, output_dir)
        else:
            print("Invalid choice. Please enter 'S' for sheets or 'F' for files.")
            logging.warning("Invalid split type choice.")
    except Exception as e:
        print(f"An error occurred while splitting the file: {e}")
        logging.error(f"An error occurred while splitting the file: {e}", exc_info=True)
    finally:
        if unprotected_file:
            os.unlink(unprotected_file)

def send_to_file(df, cols, column_name, file, sheet_name, output_dir):
    """
    Splits the DataFrame into multiple files based on unique column values.

    Parameters:
    df (DataFrame): The DataFrame to split.
    cols (array): The unique values in the column.
    column_name (str): The column name to split by.
    file (str): The original file path.
    sheet_name (str): The sheet name (if applicable).
    output_dir (str): The output directory.

    Returns:
    None
    """
    base_filename = f"{os.path.splitext(os.path.basename(file))[0]}_{sheet_name or 'Sheet1'}"
    os.makedirs(output_dir, exist_ok=True)
    for value in tqdm(cols, desc="Splitting data into files"):
        output_file = get_timestamped_filename(output_dir, f'{base_filename}_{column_name}_{value}', '.xlsx')
        filtered_df = df[df[column_name] == value]
        filtered_df.to_excel(output_file, sheet_name=str(value), engine='openpyxl', index=False)
    print('Data split into files successfully.')
    logging.info('Data split into files successfully.')

def send_to_sheet(df, cols, column_name, file, sheet_name, output_dir):
    """
    Splits the DataFrame into multiple sheets based on unique column values.

    Parameters:
    df (DataFrame): The DataFrame to split.
    cols (array): The unique values in the column.
    column_name (str): The column name to split by.
    file (str): The original file path.
    sheet_name (str): The sheet name (if applicable).
    output_dir (str): The output directory.

    Returns:
    None
    """
    output_file = get_timestamped_filename(output_dir, f'{os.path.splitext(os.path.basename(file))[0]}_{sheet_name or "Sheet1"}_split', '.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for value in tqdm(cols, desc="Splitting data into sheets"):
            sn = sanitize_sheet_name(str(value))
            filtered_df = df[df[column_name] == value]
            filtered_df.to_excel(writer, sheet_name=sn, index=False)
    print(f'Data split into sheets successfully. Output File: {output_file}')
    logging.info(f'Data split into sheets successfully. Output File: {output_file}')

def interactive_mode():
    """
    Runs the utility in interactive mode.
    """
    print_help_message()
    while True:
        user_input = input("Enter your command: ").strip()
        if user_input.lower() == 'q':
            break
        elif user_input.lower().startswith(('c ', 's ')):
            parts = user_input.strip().split(maxsplit=1)
            if len(parts) < 2:
                print("Please provide a path after the command.")
                continue
            operation, path = parts
            operation = operation.lower()
            if operation == 'c':
                combine_excel_files(path)
            elif operation == 's':
                split_excel_file(path)
            else:
                print("Invalid command. Type 'Q' to quit or see above for command usage.")
        elif user_input.lower() == 'h':
            print_help_message()
        else:
            print("Invalid command. Type 'Q' to quit or see above for command usage.")

def parse_arguments():
    """
    Parses command-line arguments.

    Returns:
    Namespace: The parsed arguments.
    """
    parser = argparse.ArgumentParser(description="Manage Excel and CSV Files", add_help=False)
    parser.add_argument('-c', '--combine', help='Path to combine files')
    parser.add_argument('-s', '--split', help='File to split into multiple sheets or files')
    parser.add_argument('-o', '--output', help='Output directory for the results')
    parser.add_argument('-p', '--password', help='Password for encrypted Excel files')
    parser.add_argument('--combine-mode', choices=['sheet', 'workbook'], default='workbook', help='Combine into one sheet or one workbook')
    parser.add_argument('--sheet-name', help='Specify the sheet name to use')
    parser.add_argument('--file-type', choices=['excel', 'csv', 'all'], default='excel', help='Specify the file type to process')
    parser.add_argument('-h', '--help', action='store_true', help='Show help message and exit')
    return parser.parse_args()

def main():
    """
    The main function that executes the utility based on the provided arguments.
    """
    args = parse_arguments()
    if args.help:
        print_help_message()
        sys.exit(0)
    if args.combine:
        combine_excel_files(
            path=args.combine,
            output_dir=args.output,
            combine_mode=args.combine_mode,
            password=args.password,
            file_type=args.file_type
        )
    elif args.split:
        split_excel_file(
            file=args.split,
            output_dir=args.output,
            password=args.password,
            sheet_name=args.sheet_name,
            file_type=args.file_type
        )
    else:
        interactive_mode()

if __name__ == "__main__":
    main()
